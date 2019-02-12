from openpyxl import load_workbook

wb = load_workbook('text.xlsx')
ws1 = wb.worksheets[0]
ws2 = wb.worksheets[1]
#print(ws['a2'].value)
# print(ws[1])
# print(ws[1][0].value)

col_1 = (list(ws1.columns))
col_2 = (list(ws2.columns))

"""
usuń pierwszy element z listy
"""
def rem_first(lis):
    temp_lis = []
    for ele in lis:
        ele = list(ele)
        ele.pop(0)
        temp_lis.append(ele)

    return temp_lis

col_1 = rem_first(col_1)
col_2 = rem_first(col_2)

"""
pobierz wartości z komurek
"""
def get_value(cols):
    temp_list = []
    for col in cols:
        values = []
        for cell in col:
            values.append(cell.value)
        temp_list.append(values)

    return temp_list

col_1 = get_value(col_1)
col_2 = get_value(col_2)


"""
połącz dwie listy
"""
fulllist = []
for a, b in zip(col_1, col_2):
    a.extend(b)
    fulllist.append(a)

ws3 = wb.create_sheet()
"""
wklej dane do arkusza
"""
for c in range(len(fulllist)):
    for r in range(len(fulllist[c])):
        ws3.cell(row=1+r, column=1+c, value=fulllist[c][r])



wb.save('text.xlsx')




