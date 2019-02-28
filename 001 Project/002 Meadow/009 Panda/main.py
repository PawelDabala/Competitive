from openpyxl import load_workbook
wb = load_workbook('channel.xlsx')
ws1 = wb.worksheets[0]

col_1 = (list(ws1.columns))

"""
usuń pierwszy element z listy
"""
def rem_first(lis):
    temp_lis = []
    for ele in lis:
        ele = list(ele)
        ele.pop(0)
        temp_lis.append(ele)

    return temp_lis

col_1 = rem_first(col_1)


"""
pobierz wartości z komurek
"""
def get_value(cols):
    temp_list = []
    for col in cols:
        values = []
        for cell in col:
            values.append(cell.value)
        temp_list.append(values)

    return temp_list

col_1 = get_value(col_1)
# print(col_1)
col_2 = list(zip(col_1[0],col_1[1]))
print(len(col_2))
print(len(set(col_2)))
col_2 = set(col_2)
ws3 = wb.create_sheet()

for nr, c, in enumerate(col_2, 1):
    for r, entry in enumerate(c, start=1):
        ws3.cell(row=r, column=nr, value=entry)


wb.save('channel.xlsx')
