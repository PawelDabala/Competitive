import sys
from PySide2.QtWidgets import *
from PySide2.QtCore import Qt
from uiexcel import Ui_Dialog

from sqlalchemy import and_
from sql.filterf import FilterF
from sql.base import Session
from sql.category import Category
from sql.competitive import Competitive
from sql.data import Data
from sql.compatitive_filter import CompativeFilterf

from openpyxl import Workbook



class ExcelForm(QDialog):
    def __init__(self, main_headers, parent=None):
        super(ExcelForm, self).__init__(parent)
        self.ui = Ui_Dialog()
        self.ui.setupUi(self)
        self.setWindowTitle('Generuj plik Excela')
        self.setWindowModality(Qt.ApplicationModal)
        self.main_headers = main_headers

        self.cb_raport_name = self.ui.comboBox_raport_name
        self.lv_year = self.ui.listWidget_year
        self.lv_month = self.ui.listWidget_mont
        self.lv_week = self.ui.listWidget_week
        self.lv_media = self.ui.listWidget_medium

        self.pb_excel = self.ui.pushButton_excel
        self.pb_cancel = self.ui.pushButton_close

        self.compative_name = ""

        #signals
        self.pb_cancel.clicked.connect(self.close)
        self.pb_excel.clicked.connect(self.generate_excel_file)
        self.cb_raport_name.currentTextChanged.connect(self.set_filter_section)

        self.lv_year.itemChanged.connect(self.lw_channge_year)
        self.lv_month.itemChanged.connect(self.lw_channge_month)


        #functions
        self.set_compatives()
        self.set_filter_section()

    def set_compatives(self):
        """
        read rows from database with competitive name
        :return:
        """
        try:
            session = Session()
            competitev = session.query(Competitive).all()

        except:
            QMessageBox.critical(self, "Błąd", "Nie można połączyć się z bazą danych")
            return

        self.cb_raport_name.clear()
        for compet in competitev:
            self.cb_raport_name.addItem(compet.name)

        session.close()

    def set_filter_section(self):
        session = Session()
        id_nr = self.get_data_id()

        year = [ye.year for ye in
                session.query(Data).distinct(Data.year).filter(Data.competitive_id == id_nr).order_by(
                    Data.year).all()]
        self.set_cb_value(self.lv_year, year)

        month = [mo.month for mo in
                 session.query(Data).distinct(Data.month).filter(Data.competitive_id == id_nr).order_by(
                     Data.month).all()]
        self.set_cb_value(self.lv_month, month)

        week = [we.week_nr for we in
                 session.query(Data).distinct(Data.week_nr).filter(Data.competitive_id == id_nr).order_by(
                     Data.week_nr).all()]
        self.set_cb_value(self.lv_week, week)

        media = [me.media for me in
                    session.query(Data).distinct(Data.media).filter(Data.competitive_id == id_nr).order_by(
                    Data.media).all()]
        self.set_cb_value(self.lv_media, media)


        session.close()

    def get_data_id(self):

        com_name = self.cb_raport_name.currentText()
        self.compative_name = com_name

        if com_name != '':
            try:
                session = Session()
                compative_id = session.query(Competitive.id).filter_by(name=com_name).one()
            except:
                QMessageBox.critical(self, "Błąd", "Nie można połączyć się z bazą danych")
                return

            session.close()
            return compative_id.id

    def set_cb_value(self, lv, value):
        """
        add value to combo box
        :param cb:
        :param value:
        :return:
        """
        lv.clear()
        for val in value:
            item = QListWidgetItem(str(val))
            item.setCheckState(Qt.Checked)
            lv.addItem(item)

    def set_table(self):
        """
        set tabel with data from data base
        :return:
        """

        #temporaty get all data, after change to set query
        id_nr = self.get_data_id()
        years = self.get_checked_items(self.lv_year)
        months = self.get_checked_items(self.lv_month)
        weeks = self.get_checked_items(self.lv_week)
        media = self.get_checked_items(self.lv_media)
        session = Session()

        # data = session.query(Data).filter_by(competitive_id=id_nr).all()
        data = session.query(Data).filter(and_(
            Data.year.in_(years),
            Data.month.in_(months),
            Data.week_nr.in_(weeks),
            Data.media.in_(media),
            Data.competitive_id == id_nr,
        )).all()
        headers = Data.__table__.columns.keys()
        #remove id, compatitive_id
        headers = headers[2:]
        #make list
        final_list = []
        for row in data:
            columns = []
            for key in headers:
                columns.append(row.__dict__[key])
            final_list.append(columns)

        session.close()
        return final_list

    def set_excel_file(self, final_list):
        """
        preper excel file from final_list and self.main_headers
        :param final_list:
        :return:
        """
        wb = Workbook()
        ws = wb.active

        #set heders
        for c, c_val in enumerate(self.main_headers, 1):
            ws.cell(row=1, column=c, value=c_val)

        #set data to excel file
        for r, row in enumerate(final_list, 2):
            for c, entry in enumerate(row, 1):
                ws.cell(row=r, column=c, value=entry)
        name = QFileDialog.getSaveFileName(self, caption="Zapisz",filter='.xlsx',selectedFilter='.xlsx')
        name = ''.join(name)
        wb.save(name)
        QMessageBox.information(self, "Zapis", "Plik został utworzony")


    def generate_excel_file(self):
        """
        Make execel file
        :return:
        """
        final_list = self.set_table()
        self.set_excel_file(final_list)


    """
    
    List view section
    
    """
    def lw_channge_year(self):

        years = self.get_checked_items(self.lv_year)

        session = Session()
        compative = session.query(Competitive).filter_by(name=self.compative_name).one()

        #month
        month = [mo.month for mo in
                 session.query(Data).distinct(Data.month).filter(and_(
                     Data.year.in_(years),
                     Data.competitive_id == compative.id)).order_by(
                     Data.month).all()]
        self.set_cb_value(self.lv_month, month)

        #week
        week = [we.week_nr for we in
                session.query(Data).distinct(Data.week_nr).filter(and_(
                    Data.year.in_(years),
                    Data.month.in_(month),
                    Data.competitive_id == compative.id,
                    )).order_by(
                    Data.week_nr).all()]
        self.set_cb_value(self.lv_week, week)

        #media
        media = [me.media for me in
                session.query(Data).distinct(Data.media).filter(and_(
                    Data.year.in_(years),
                    Data.month.in_(month),
                    Data.week_nr.in_(week),
                    Data.competitive_id == compative.id,
                )).order_by(
                    Data.media).all()]
        self.set_cb_value(self.lv_media, media)

        session.close()

    def lw_channge_month(self):

        years = self.get_checked_items(self.lv_year)
        months = self.get_checked_items(self.lv_month)

        session = Session()
        compative = session.query(Competitive).filter_by(name=self.compative_name).one()

        # week
        week = [we.week_nr for we in
                session.query(Data).distinct(Data.week_nr).filter(and_(
                    Data.year.in_(years),
                    Data.month.in_(months),
                    Data.competitive_id == compative.id,
                )).order_by(
                    Data.week_nr).all()]
        self.set_cb_value(self.lv_week, week)

        media = [me.media for me in
                 session.query(Data).distinct(Data.media).filter(and_(
                     Data.year.in_(years),
                     Data.month.in_(months),
                     Data.week_nr.in_(week),
                     Data.competitive_id == compative.id,
                 )).order_by(
                     Data.media).all()]
        self.set_cb_value(self.lv_media, media)

        session.close()

    def lw_change_media(self):
        years = self.get_checked_items(self.lv_year)
        months = self.get_checked_items(self.lv_month)
        week = self.get_checked_items(self.lv_week)

        session = Session()
        compative = session.query(Competitive).filter_by(name=self.compative_name).one()

        media = [me.media for me in
                 session.query(Data).distinct(Data.media).filter(and_(
                     Data.year.in_(years),
                     Data.month.in_(months),
                     Data.week_nr.in_(week),
                     Data.competitive_id == compative.id,
                 )).order_by(
                     Data.media).all()]
        self.set_cb_value(self.lv_media, media)


    def get_checked_items(self, lw):
        """
        get checked years from listWidget
        :return:
        """
        items = []
        for nr in range(lw.count()):
            if lw.item(nr).checkState() == Qt.Checked:
                try:
                    items.append(int(lw.item(nr).text()))
                except ValueError:
                    items.append(lw.item(nr).text())

        return items


if __name__ == '__main__':
    app = QApplication(sys.argv)
    app.setStyle('fusion')
    w = ExcelForm()
    w.show()
    sys.exit(app.exec_())

